from flask import Flask, request, jsonify, session
from flask import render_template
# from gevent.pywsgi import WSGIServer
try:
    from SimConnect import *
except ImportError:
    pass

import subprocess
import multiprocessing
import os, psutil
import logging
import math, random, time
import socket, struct, pickle
import json
from openpyxl import Workbook, load_workbook

app = Flask(__name__)

# Creating simconnection
try:
    sm = SimConnect()
    aq = AircraftRequests(sm, _time=10)
except:
    sm = None
    aq = None

ships_process = None  # the C++ exe for spawning the ships
# init center surveillance area at 33.0, -118.0
shift_north = 0.00
shift_west = 0.02
scale = 86.4* 1/10 # lat long scale of the ship group
dest_lat = None
dest_long = None
old_lat = 0
old_long = 0
# init center surveillance area at 33.0, -118.0
# aircraft position
ac_init = True
ac_lat = 33.0
ac_long = -118.0
ac_old_lat = 33.0
ac_old_long = -118.0
ac_alt = 2000
ac_heading = 0.01
ac_old_heading = 0.01
ac_Xpx = 0.0
ac_Ypx = 0.0
ac_headingPx = 0.0
# gameWidth is 100NM ~ 1.44 lat long degrees 
# 1.44 degree * 60 min/degree = 86.4 minutes 
ac_scale = 86.4 * 1/10 # scale down to 10Nm real world width
allMouseData = []
phaseData = []
clickData = []
alertData = []
targetData= []
random.seed(time.time())
rand = random.randint(1000, 9999)
app.id = None

study_scenario = 1
study_sequence = 1
study_config = "none"
study_survey = "none"

mission_complete = False

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/control")
def control():
    return render_template("controller.html")

def pxToLat(pxY):
    return 33 - ((pxY-0.5) * ac_scale) / 60.0 

def pxToLong(pxX):
    return -118 + ((pxX-0.5) * ac_scale) / 60.0 

def latToPx(latY):
    return 0.5 - (latY - 33) * 60.0 / ac_scale

def longToPx(longX):
    return (((longX + 118) * 60.0 )/ ac_scale) + 0.5

@app.route("/aircraft", methods=["GET"])
def update_aircraft():
    # data = request.get_json()
    # lat = 32 + (40 - (data.get('y')-.5) * scale) / 60.0 - shift_west
    # long = -117 - (13.2 - (data.get('x')-.5) * scale) / 60.0 - shift_north
    # alt = data.get('altitude')

    # global old_lat, old_long
    # heading = -1 * math.atan2(lat - old_lat, long-old_long) + 3.14159/2 # data.get('heading')
    # old_lat = lat
    # old_long = long
    global ac_init, ac_lat, ac_long, ac_heading, ac_old_lat, ac_old_long, ac_old_heading, sm, aq
    # initialization hover
    if (ac_init):
        aq.set("PLANE_LATITUDE", ac_lat)
        aq.set("PLANE_LONGITUDE", ac_long)
        aq.set("PLANE_ALTITUDE", ac_alt)
        aq.set("PLANE_HEADING_DEGREES_TRUE", ac_heading)#data.get("heading"))
        aq.set("PLANE_PITCH_DEGREES", 0.0)
        aq.set("PLANE_BANK_DEGREES", 0.0)
        
   
    # temp fix to keep a/c upright    
    #aq.set("PLANE_ALTITUDE", ac_alt)
    #aq.set("PLANE_PITCH_DEGREES", 0.0)
    #aq.set("PLANE_BANK_DEGREES", 0.0)
    # remove temp stabilization when MATLAB implements
    # sometimes SimConnect breaks and throws an OS Error, so we are saving the current lat/long when it works (or sending the last one)
    try:
        ac_lat = aq.get("PLANE_LATITUDE")
        ac_long = aq.get("PLANE_LONGITUDE")
        ac_heading = aq.get("PLANE_HEADING_DEGREES_TRUE")
    except:
        logging.info("SimConnect Error in /var")
        #print("\n SimConnect Error in /var")
        try:
            sm = SimConnect()
            aq = AircraftRequests(sm, _time=10)
        except:
            pass
        pass

    if (ac_lat != None): ac_old_lat = ac_lat
    if (ac_long != None): ac_old_long = ac_long
    if (ac_heading != None): ac_old_heading = ac_heading

    #print("\n MSFS in a/c lat:", ac_old_lat, " long:", ac_old_long, " heading:", ac_old_heading)

    ac_Xpx = longToPx(ac_old_long)
    ac_Ypx = latToPx(ac_old_lat)
    ac_headingPx = float(ac_old_heading)

    #print("\n a/c X:", ac_Xpx, " Y", ac_Ypx, " Heading:", ac_headingPx* 180/3.1459)
    return_dict = {"x": ac_Xpx,
                   "y": ac_Ypx,
                   "heading": ac_headingPx}

    return jsonify(return_dict) 

# spawns the ships given waypoints
@app.route("/ships", methods=['POST'])
def spawn_ships():
    # assumes the 'ships' param is structured as:
    #
    # [
    #   [  # ship
    #     [ID, targetClass, threatClass]
    #     [x, y, s],  # waypoint
    #     [x, y, s],
    #     [x, y, s],
    #     ...
    #   ],
    #   ...
    # ]
    
    data = request.get_json()
    ships = data.get('ships')
    numShips = len(ships)
    shipstring = ""

    # create blank array with MATLABarraysize=10 ship elements
    matlabObsArrSize = 20
    shipObstacles = []
    shipID = [0.0 for _ in range(matlabObsArrSize)]
    shipTargetClass = [0.0 for _ in range(matlabObsArrSize)]
    shipThreatClass = [0.0 for _ in range(matlabObsArrSize)]
    shipLong = [0.0 for _ in range(matlabObsArrSize)]
    shipLat = [0.0 for _ in range(matlabObsArrSize)]

    for i in range(numShips):
        ship = ships[i]
        shipID[i % matlabObsArrSize] =  float(ship[0][0])
        shipTargetClass[i % matlabObsArrSize]  = float(ship[0][1])
        shipThreatClass[i % matlabObsArrSize]  = float(ship[0][2])
        shipLong[i % matlabObsArrSize]  = pxToLong(ship[1][0])
        shipLat[i % matlabObsArrSize]  = pxToLat(ship[1][1])
        for waypoint in ship[1:]:
            shipstring += str(waypoint[0]) + "-" + str(waypoint[1]) + "-" + str(waypoint[2]) + ","
        shipstring += ";"
    
    with open("./waypoints.txt", "w") as f:
        f.write(str(shift_north) + "\n")  # shift north (slight changes should be on the order of 0.01)
        f.write(str(shift_west) + "\n")  # shift west (slight changes should be on the order of 0.01)
        f.write(str(scale) + "\n")  # lat long scale (group size)
        f.write(shipstring)

    # send ship info and spawn point to matlab for CBF obstacle creation
    shipObstacles = [shipID, shipLat, shipLong, shipTargetClass, shipThreatClass]
    return_status = matlab_update_ship(shipObstacles)
    print(return_status)

    global ships_process
    # delete the ships if they are already active
    if ships_process is not None:
        # the generated process spawns the AIObjects.exe, which is a separate process, so we kill all children of the spawned process
        # something simple like process.terminate() does not kill AIObjects.exe
        parent = psutil.Process(ships_process.pid)
        for child in parent.children(recursive=True):  # or parent.children() for recursive=False
            child.kill()
        parent.kill()      

    # create the process
    ships_process = multiprocessing.Process(target=start_ships, args=(shipstring,))
    ships_process.start()
    return ""

def start_ships(shipstring=""):
    with open("./waypoints.txt", "rb") as f:
        subprocess.run([".\AIObjects.exe"], stdin=f, text=True, shell=False)
    print("\n done!")
    return
    
def matlab_update_ship(shipObs):
    MATLAB_IP = '127.0.0.1'
    MATLAB_PORT_OBS_ID = 8082
    MATLAB_PORT_OBS_LAT = 8084
    MATLAB_PORT_OBS_LONG = 8086
    MATLAB_PORT_OBS_TARGET_CLASS = 7088
    MATLAB_PORT_OBS_THREAT_CLASS = 7080
    

    
    # format obstacle info
    IDs = shipObs[0]
    ship_lats = shipObs[1]
    ship_longs = shipObs[2]
    ship_target_class = shipObs[3]
    ship_threat_class = shipObs[4]

    # debug 
    print(shipObs)

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)

    for i in range(len(IDs)):
        # send the ship obstacle info
        s.sendto(struct.pack('>d', IDs[i]), (MATLAB_IP, MATLAB_PORT_OBS_ID ))
        s.sendto(struct.pack('>d', ship_lats[i]), (MATLAB_IP, MATLAB_PORT_OBS_LAT))
        s.sendto(struct.pack('>d', ship_longs[i]), (MATLAB_IP, MATLAB_PORT_OBS_LONG))
        s.sendto(struct.pack('>d', ship_target_class[i]), (MATLAB_IP, MATLAB_PORT_OBS_TARGET_CLASS ))
        s.sendto(struct.pack('>d', ship_threat_class[i]), (MATLAB_IP, MATLAB_PORT_OBS_THREAT_CLASS ))

    return "success"

@app.route('/run-executable', methods=['POST'])
def run_executable():
    dataArr = request.get_json()
    
    for data in dataArr:
        id = data.get('id')
        tClass = data.get('class')
        speed = data.get('speed')
        x = data.get('x')
        y = data.get('y')
        wpx = data.get('wpx')
        wpy = data.get('wpy')
        print(f"ID: {id}. Class: {tClass}. x,y: ({x},{y}). wpx,wpy: ({wpx},{wpy}). Speed: {speed}")
    return "success"

# clientside configuration route, client reads this to determine what configuration to show
@app.route("/config", methods=["GET"])
def get_config():
    return jsonify({"config": study_config, "scenario": study_scenario, "sequence": study_sequence, "survey": study_survey})

@app.route("/config_ISR1", methods=["POST"])
def post_config_ISR1():
    data = request.get_json()

    survey = data.get("survey")
    sequence = int(data.get("sequence"))
    scenario = int(data.get("scenario"))

    square_order = ["A0", "B0", "B3", "A1", "A3", "B1", "B2", "A2",
                    "B0", "A1", "A0", "B1", "B3", "A2", "A3", "B2",
                    "A1", "B1", "B0", "A2", "A0", "B2", "B3", "A3",
                    "B1", "A2", "A1", "B2", "B0", "A3", "A0", "B3",
                    "A2", "B2", "B1", "A3", "A1", "B3", "B0", "A0",
                    "B2", "A3", "A2", "B3", "B1", "A0", "A1", "B0",
                    "A3", "B3", "B2", "A0", "A2", "B0", "B1", "A1",
                    "B3", "A0", "A3", "B0", "B2", "A1", "A2", "B1"]
    if sequence > 8:
        return "sequence too high, capped at 8"
    if scenario > 8:
        return "scenario too high, capped at 8"
    if sequence < 1:
        return "sequence too low, min is 1"
    if scenario < 1:
        return "scenario too low, min is 1"
    if sequence * scenario > len(square_order):
        return "sequence x scenario is greater than max: " + str(len(square_order))
    global study_scenario, study_sequence, study_config, study_survey

    study_config = square_order[((sequence-1) * 8) + scenario - 1]
    study_scenario = scenario
    study_sequence = sequence
    study_survey = survey
    return "Server: updated scenario (" + str(scenario) + ") and sequence (" + str(sequence) + "), new config: " + str(study_config) + ", last survey: " + str(survey)

@app.route("/config", methods=["POST"])
def post_config():
    data = request.get_json()

    survey = data.get("survey")
    sequence = int(data.get("sequence"))
    scenario = int(data.get("scenario"))

    square_order = ["A0", "B0", "B4", "A2", "A4", "B2",
                    "B0", "A2", "A0", "B2", "B4", "A4", 
                    "A2", "B2", "B0", "A4", "A0", "B4", 
                    "B2", "A4", "A2", "B4", "B0", "A0", 
                    "A4", "B4", "B2", "A0", "A2", "B0", 
                    "B4", "A0", "A4", "B0", "B2", "A2"]
    if sequence > 6:
        return "sequence too high, capped at 8"
    if scenario > 6:
        return "scenario too high, capped at 8"
    if sequence < 1:
        return "sequence too low, min is 1"
    if scenario < 1:
        return "scenario too low, min is 1"
    if sequence * scenario > len(square_order):
        return "sequence x scenario is greater than max: " + str(len(square_order))
    global study_scenario, study_sequence, study_config, study_survey

    study_config = square_order[((sequence-1) * 6) + scenario - 1]
    study_scenario = scenario
    study_sequence = sequence
    study_survey = survey

    # aircraft position
    global ac_lat, ac_long, ac_old_lat, ac_old_long, ac_heading, ac_old_heading, ac_Xpx, ac_Ypx, ac_headingPx
    ac_lat = 33.0
    ac_long = -118.0
    ac_old_lat = 33.0
    ac_old_long = -118.0
    ac_heading = 0.01
    ac_old_heading = 0.01
    ac_Xpx = 0.0
    ac_Ypx = 0.0
    ac_headingPx = 0.0

    return "Server: updated scenario (" + str(scenario) + ") and sequence (" + str(sequence) + "), new config: " + str(study_config) + ", last survey: " + str(survey)

# logging route
@app.route("/log", methods=["POST"])
def log():
    #print("logging!", request.get_json())
    return ""

log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

@app.route("/saveId", methods=["POST"])
def saveId():
    data = request.get_json()
    app.id = data.get('userId')
    print(app.id)
    return ""

@app.route("/saveData", methods=["POST"])
def saveData():
    data = request.get_json()
    newData = {
        "phase": data.get('phase'),
        "score": data.get('score'),
        "numClicks": data.get('numClicks'),
        "time": data.get('time'),
        "unixTime": data.get('unixTime'),
        "denied": data.get('denied'),
        "numOptimized": data.get('numOptimized'),
        "wez_alerts": data.get('wez_alerts'),
        "tmv_alerts": data.get('tmv_alerts'),
        "osi_alerts": data.get('osi_alerts'),
        "utc_alerts": data.get('utc_alerts')
    }
    phaseData.append(newData)
    filename = f"{app.id}_{study_config}.pkl"
    txtFilename = f"{app.id}_{study_config}.txt"

    with open(filename, 'wb') as file:
        pickle.dump(allMouseData, file)
    with open(filename, 'ab') as file:
        pickle.dump(phaseData, file)
    with open(filename, 'ab') as file:
        pickle.dump(clickData, file)
    with open(filename, 'ab') as file:
        pickle.dump(alertData, file)
    with open(filename, 'ab') as file:
        pickle.dump(targetData, file)
    text_content = "\n".join(json.dumps(element) if isinstance(element, dict) else element for element in alertData)
    with open(txtFilename,"w") as file:
        file.write(text_content)


    #reset data structures at current mission end
    if (data.get('phase') == 2 and mission_complete == True):
        exFilename = f"{app.id}_{study_config}_alerts.xlsx"
        if not os.path.isfile(exFilename):
            wb = Workbook()
            ws = wb.active
            ws.append(['User ID'] + ['WEZ'] + ['Optimized Search'] + ['Unable to Comply'] + ['Too Many Vectors'] 
                      + ['WEZ'] + ['Optimized Search'] + ['Unable to Comply'] + ['Too Many Vectors'])
            wb.save(exFilename)
        wb = load_workbook(exFilename)
        ws = wb.active
        ws.append([app.id] + [phaseData[0]['wez_alerts']] + [phaseData[0]['osi_alerts']] + [phaseData[0]['utc_alerts']] + [phaseData[0]['tmv_alerts']]
                  + [phaseData[1]['wez_alerts']] + [phaseData[1]['osi_alerts']] + [phaseData[1]['utc_alerts']] + [phaseData[1]['tmv_alerts']])
        wb.save(exFilename)
        phaseData.clear()
        clickData.clear()
        allMouseData.clear()
        alertData.clear()
        targetData.clear()
    return ""

@app.route('/receive-mouse-coordinates', methods=["POST"])
def receive_mouse_coordinates():
    data = request.json
    allMouseData.extend(data)
    return jsonify({})

@app.route('/receive-mouse-click', methods=["POST"])
def receive_mouse_click():
    data = request.json
    clickData.extend(data)
    return jsonify({})

@app.route('/receive-target-data', methods=["POST"])
def receive_target_data():
    data = request.json
    targetData.extend(data)
    return jsonify({})

@app.route('/receive-alert', methods=["POST"])
def receive_alert():
    data = request.json
    alertData.extend(data)
    return jsonify({})

@app.route('/set-mission-complete', methods=["POST"])
def set_mission_complete():
    global mission_complete
    mission_complete = request.json.get("missionEnd", False)
    print("Set mission complete to ", mission_complete)
    return jsonify({})

@app.route('/get-mission-complete', methods=["GET"])
def get_mission_complete():
    #print("Sending mission complete ", mission_complete)
    return jsonify({"missionEnd": mission_complete})

# data for MATLAB route
@app.route("/current-destination", methods=["POST"])
# UDP update for MATLAB
def matlab_destination_update():
    MATLAB_IP = '127.0.0.1'
    MATLAB_PORT_LAT_MIN = 8080
    MATLAB_PORT_LONG_MIN = 8088

    global dest_lat, dest_long, ac_init
    # on first destination, release aircraft from init hover
    ac_init = False
    data = request.get_json()
    dest_x = float(data.get('dest x'))
    dest_y = float(data.get('dest y'))
    dest_long = pxToLong(dest_x )  # convert to long
    dest_lat = pxToLat(dest_y)  # convert to lat
 

    if (dest_lat is None or dest_long is None):
        return "failure"

    # parse the latitude and longitude
    latitude = float(dest_lat)
    longitude = float(dest_long)

    #print("in a/c dest x:", dest_x,  " y:", dest_y )
    print("out a/c dest lat:", latitude, " long:", longitude)

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)

    # send the location 1x times to mitigate UDP packet drops
    for _ in range(1):
        s.sendto(struct.pack('>f', latitude), (MATLAB_IP, MATLAB_PORT_LAT_MIN))
        s.sendto(struct.pack('>f', longitude), (MATLAB_IP, MATLAB_PORT_LONG_MIN))

    return "success"

@app.route("/cognitivestate", methods=["POST"])
def matlab_cognitivestate():
    MATLAB_IP = '127.0.0.1'
    MATLAB_PORT_COG_STATE = 7082

    data = request.get_json()

    data_cogstate = data.get("cogstate")
    cogstate = float(data_cogstate)
    print(cogstate)

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.sendto(struct.pack('>d', cogstate), (MATLAB_IP, MATLAB_PORT_COG_STATE))

    return "success"

@app.route("/attentionstate", methods=["POST"])
def matlab_attentionstate():
    #print("we're calling /attentionstate")
    LAOIs = request.get_json()["LAOIs"]
    #print(LAOIs, type(LAOIs))
    MATLAB_IP = '127.0.0.1'
    MATLAB_PORT_ATT_STATE = 7082
    #Getting 1st LAOI state
    attstate = float((LAOIs >> 0) & 1)

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.sendto(struct.pack('>d', attstate), (MATLAB_IP, MATLAB_PORT_ATT_STATE))

    return "success"

@app.route("/radio")
def radio_panel():
    return render_template("radio_panel.html")

if __name__ == '__main__':
    app.run("0.0.0.0", port=100, debug=False)
